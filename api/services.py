import io
import json, requests
import pandas as pd
from openpyxl import load_workbook


def get_market_rate(part):
    response = requests.post('https://mosapi.sinewave.co.in/stocks/', json={'tickers': part}, verify=False)
    string = response.json()
    json_res = json.loads(string)
    if len(json_res) > 0:
        data = json_res[0]
        return data
    else:
        return None


def get_market_rate_value(part):
    data = get_market_rate(part)

    if data:
        return data['Adj Close']
    else:
        return None


def get_strategy(parts, days):
    response = requests.post('http://filerenderapi.sinewave.co.in/api/', json={'tickers': parts, 'per': int(days)},
                             verify=False)
    string = response.content
    toread = io.BytesIO(string)  # pass your `decrypted` string as the argument here
    toread.seek(0)  # reset the pointer
    return toread


def get_strategy_values(parts, days):
    toread = get_strategy(parts, days)
    result = []
    wb = load_workbook(toread, read_only=True)  # open an Excel file and return a workbook
    for part in parts:
        if part not in wb.sheetnames:
            continue
        df = pd.read_excel(toread, sheet_name=part)
        last = df.tail(1).iloc[0]
        record = {
            'part': part,
            'date': last['Date'],
            'trigger': last['ATR']
        }
        result.append(record)

    return result
